import csv
from jobspy import scrape_jobs
import openpyxl
import os
#import openai
#from docx import Document

#import airflow 

jobs = scrape_jobs(
    site_name=["linkedin", "google"],
    search_term="Data Scientist", 
    google_search_term="SQL or Data Scientist jobs near Toronto, ON since yesterday",
    location="Toronto, ON",
    results_wanted=30,
    hours_old=72,
    country_indeed='canada',
    linkedin_fetch_description=True, # gets more info such as description, direct job url (slower)
    #proxies=["208.195.175.46:65095", "208.195.175.45:65095", "localhost"],
)
print(f"Found {len(jobs)} jobs")
print(jobs.head())
file_path = "jobs.xlsx"

# Check if the file exists
if os.path.exists(file_path):
    workbook = openpyxl.load_workbook(file_path)
    if "Sheet" in workbook.sheetnames:
        sheet = workbook["Sheet"]
    else:
        sheet = workbook.create_sheet("Sheet")
else:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sheet"
    # Write the header if the file is new
    header = list(jobs.columns)
    sheet.append(header)

# Find the next empty row
next_row = sheet.max_row + 1

# Write the job data
for index, row in jobs.iterrows():
    for col_num, value in enumerate(row, start=1):
        sheet.cell(row=next_row, column=col_num, value=value)
    next_row += 1

# Save the workbook
workbook.save(file_path)

# Ensure the file is not deleted when closed
workbook.close()


# Set up the OpenAI API client
# openai.api_key = 'your_openai_api_key'

# Define the prompt for generating the resume
# prompt = """
# Generate a resume for a Data Scientist based on the following job listings:
# {}
# Use the following template:
# [Your Resume Template Here]
# """

# Format the job listings into the prompt
# job_listings = jobs.to_string(index=False)
# formatted_prompt = prompt.format(job_listings)

# Generate the resume using the OpenAI API
# response = openai.Completion.create(
#     engine="text-davinci-003",
#     prompt=formatted_prompt,
#     max_tokens=1500
# )

# Get the generated resume text
# resume_text = response.choices[0].text.strip()

# Load the existing document template
# template_path = "resume_template.docx"
# doc = Document(template_path)

# Find the section to edit (assuming a placeholder text like "[Job Listings]" exists)
# for paragraph in doc.paragraphs:
#     if "[Job Listings]" in paragraph.text:
#         paragraph.text = resume_text
#         break

# Save the modified document
# doc.save("generated_resume.docx")
